import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def generate_report(session_id):

    os.makedirs("reports", exist_ok=True)

    conn = sqlite3.connect("database/attendance.db")

    attendance = pd.read_sql_query(
        "SELECT roll, name, ip FROM attendance WHERE session_id=?",
        conn,
        params=(session_id,)
    )

    conn.close()

    # Example class roster
    roster = pd.DataFrame({
        "Roll Number": [str(i) for i in range(1, 61)]  # 1–60 students
    })

    attendance["roll"] = attendance["roll"].astype(str)

    roster["Name"] = roster["Roll Number"].map(
        attendance.set_index("roll")["name"]
    )

    roster["IP Address"] = roster["Roll Number"].map(
        attendance.set_index("roll")["ip"]
    )

    roster["Status"] = roster["Roll Number"].apply(
        lambda r: "Present" if r in attendance["roll"].values else "Absent"
    )

    # Add metadata columns
    roster["Section"] = ""
    roster["Subject Code"] = ""
    roster["Department"] = ""

    # Reorder columns
    roster = roster[[
        "Roll Number",
        "Name",
        "Section",
        "Subject Code",
        "Department",
        "IP Address",
        "Status"
    ]]

    file_path = f"reports/attendance_{session_id}.xlsx"

    roster.to_excel(file_path, index=False)

    # Highlight absentees
    wb = load_workbook(file_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        if ws[f"G{row}"].value == "Absent":
            ws[f"G{row}"].fill = red_fill

    wb.save(file_path)

    return file_path
